#!/usr/bin/env python3
"""
Script to update HeatMap Sheet Status column based on Evaluation Results.

This script reads evaluation results from the "Evaluation Results" sheet and updates
the Status column in the "HeatMap Sheet" according to the following rules:

1. For sub-operations: Fill with colored dots (● with red/green/yellow)
2. For parent operations: Fill with text "NOK" (red), "acceptable" (yellow), "OK" (green)
3. Parent operation status is calculated based on the worst status of its sub-operations

Usage:
    python3 update_heatmap_status.py

Requirements:
    - openpyxl: pip install openpyxl
    - Excel file: AVLDrive_Heatmap_Tool version3.2.xlsm
"""

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
from collections import defaultdict
import sys
import os

# Color definitions
RED_COLOR = "FF0000"
GREEN_COLOR = "00FF00"
YELLOW_COLOR = "FFFF00"

# Column indices (0-based)
OPCODE_COLUMN = 0           # Column A
OPERATION_COLUMN = 1        # Column B
FINAL_STATUS_COLUMN = 11    # Column L (in Evaluation Results sheet)
STATUS_COLUMN = 17          # Column R (in HeatMap Sheet)

def is_parent_operation(opcode):
    """
    Determine if an OpCode is a parent operation.
    Parent operations have OpCodes ending with multiple zeros (4 or more).
    
    Args:
        opcode: The OpCode to check
        
    Returns:
        bool: True if the OpCode is a parent operation
    """
    opcode_str = str(opcode)
    if not opcode_str.isdigit():
        return False
    # Count trailing zeros - parent operations have 4 or more trailing zeros
    trailing_zeros = len(opcode_str) - len(opcode_str.rstrip('0'))
    return trailing_zeros >= 4

def get_parent_opcode(opcode_str):
    """
    Get the parent OpCode for a given sub-operation OpCode.
    The parent is the first 4 digits followed by 0000.
    
    Args:
        opcode_str: The sub-operation OpCode
        
    Returns:
        str: The parent OpCode or None if not applicable
    """
    if not opcode_str.isdigit() or len(opcode_str) < 8:
        return None
    return opcode_str[:4] + "0000"

def get_color_for_status(status):
    """
    Get color code for a given status.
    
    Args:
        status: Status string (RED, GREEN, YELLOW)
        
    Returns:
        str: Hex color code or None
    """
    if status == "RED":
        return RED_COLOR
    elif status == "GREEN":
        return GREEN_COLOR
    elif status == "YELLOW":
        return YELLOW_COLOR
    return None

def get_text_for_status(status):
    """
    Get text representation for parent operation status.
    
    Args:
        status: Status string (RED, GREEN, YELLOW)
        
    Returns:
        str: Text representation ("NOK", "OK", "acceptable") or None
    """
    if status == "RED":
        return "NOK"
    elif status == "GREEN":
        return "OK"
    elif status == "YELLOW":
        return "acceptable"
    return None

def get_worst_status(statuses):
    """
    Determine the worst status from a list of statuses.
    Priority: RED > YELLOW > GREEN > N/A
    
    Args:
        statuses: List of status strings
        
    Returns:
        str: The worst status from the list
    """
    if "RED" in statuses:
        return "RED"
    elif "YELLOW" in statuses:
        return "YELLOW"
    elif "GREEN" in statuses:
        return "GREEN"
    return "N/A"

def normalize_opcode(opcode):
    """
    Normalize an OpCode to a standard string format.
    
    Args:
        opcode: The OpCode value (can be int, float, or string)
        
    Returns:
        str: Normalized OpCode string or None
    """
    try:
        if opcode is None:
            return None
        return str(int(float(opcode)))
    except (ValueError, TypeError):
        return None

def main():
    """Main function to update the HeatMap sheet."""
    
    # File paths
    script_dir = os.path.dirname(os.path.abspath(__file__))
    input_file = os.path.join(script_dir, 'AVLDrive_Heatmap_Tool version3.2.xlsm')
    
    if not os.path.exists(input_file):
        print(f"Error: File not found: {input_file}")
        sys.exit(1)
    
    print("=" * 80)
    print("HeatMap Status Update Script")
    print("=" * 80)
    print(f"Input file: {input_file}")
    
    # Load workbook
    print("\nLoading workbook...")
    try:
        wb = load_workbook(input_file, keep_vba=True)
    except Exception as e:
        print(f"Error loading workbook: {e}")
        sys.exit(1)
    
    # Read Evaluation Results
    print("Reading Evaluation Results sheet...")
    try:
        eval_sheet = wb["Evaluation Results"]
    except KeyError:
        print("Error: 'Evaluation Results' sheet not found")
        wb.close()
        sys.exit(1)
    
    # Dictionary to store evaluation results by OpCode
    eval_results = defaultdict(list)
    
    for row in eval_sheet.iter_rows(min_row=2, max_row=eval_sheet.max_row, values_only=True):
        opcode = normalize_opcode(row[OPCODE_COLUMN])
        operation = row[OPERATION_COLUMN]
        final_status = row[FINAL_STATUS_COLUMN]
        
        if opcode and final_status and final_status not in ["N/A", None]:
            eval_results[opcode].append({
                'operation': operation,
                'status': final_status
            })
    
    print(f"Found {len(eval_results)} unique OpCodes with evaluation results")
    
    # Update HeatMap Sheet
    print("\nUpdating HeatMap Sheet...")
    try:
        heatmap_sheet = wb["HeatMap Sheet"]
    except KeyError:
        print("Error: 'HeatMap Sheet' not found")
        wb.close()
        sys.exit(1)
    
    # Track parent operations and their sub-operation statuses
    parent_operations = {}  # row_idx -> opcode
    sub_operations_by_parent = defaultdict(list)  # parent_opcode -> list of statuses
    
    # First pass: Update sub-operations and collect parent info
    sub_op_count = 0
    for row_idx in range(4, heatmap_sheet.max_row + 1):
        row = list(heatmap_sheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=False))[0]
        opcode_cell = row[OPCODE_COLUMN]
        opcode = normalize_opcode(opcode_cell.value)
        status_cell = row[STATUS_COLUMN]
        
        if not opcode:
            continue
        
        # Check if this is a parent or sub-operation
        if is_parent_operation(opcode):
            parent_operations[row_idx] = opcode
        else:
            # This is a sub-operation
            if opcode in eval_results:
                # Get the worst status from all matching evaluations
                statuses = [er['status'] for er in eval_results[opcode]]
                status = get_worst_status(statuses)
                
                if status != "N/A":
                    # Fill with colored dot
                    status_cell.value = "●"
                    color = get_color_for_status(status)
                    if color:
                        status_cell.font = Font(color=color, size=14)
                    sub_op_count += 1
                    
                    # Track status for parent calculation
                    parent_opcode = get_parent_opcode(opcode)
                    if parent_opcode:
                        sub_operations_by_parent[parent_opcode].append(status)
    
    print(f"Updated {sub_op_count} sub-operations with colored dots")
    
    # Second pass: Update parent operations based on sub-operations
    parent_op_count = 0
    for row_idx, parent_opcode in parent_operations.items():
        row = list(heatmap_sheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=False))[0]
        status_cell = row[STATUS_COLUMN]
        
        # Get all sub-operation statuses for this parent
        sub_statuses = sub_operations_by_parent.get(parent_opcode, [])
        
        if sub_statuses:
            # Calculate worst status
            worst_status = get_worst_status(sub_statuses)
            status_text = get_text_for_status(worst_status)
            
            if status_text:
                status_cell.value = status_text
                color = get_color_for_status(worst_status)
                if color:
                    status_cell.font = Font(color=color, bold=True, size=11)
                parent_op_count += 1
        else:
            # Check if the parent itself has an evaluation
            if parent_opcode in eval_results:
                statuses = [er['status'] for er in eval_results[parent_opcode]]
                status = get_worst_status(statuses)
                
                if status != "N/A":
                    status_text = get_text_for_status(status)
                    if status_text:
                        status_cell.value = status_text
                        color = get_color_for_status(status)
                        if color:
                            status_cell.font = Font(color=color, bold=True, size=11)
                        parent_op_count += 1
    
    print(f"Updated {parent_op_count} parent operations with status text")
    
    # Save the workbook
    print("\nSaving workbook...")
    try:
        wb.save(input_file)
        print(f"Successfully updated: {input_file}")
        print("\n" + "=" * 80)
        print("Update completed successfully!")
        print("=" * 80)
    except Exception as e:
        print(f"Error saving workbook: {e}")
        sys.exit(1)
    finally:
        wb.close()

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\nOperation cancelled by user")
        sys.exit(1)
    except Exception as e:
        print(f"\nUnexpected error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
