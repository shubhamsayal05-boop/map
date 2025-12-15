#!/usr/bin/env python3
"""
AVLDrive HeatMap Updater - GUI Application with AI Features

A robust, user-friendly application for updating HeatMap sheets with evaluation results.
Features include:
- Interactive GUI with file selection
- Real-time progress tracking
- AI-powered validation and recommendations
- Visual feedback and error handling
- Export and reporting capabilities
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
from collections import defaultdict
import sys
import os
import threading
from datetime import datetime
import json

# Color definitions
RED_COLOR = "FF0000"
GREEN_COLOR = "00FF00"
YELLOW_COLOR = "FFFF00"

# Column indices (0-based)
OPCODE_COLUMN = 0
OPERATION_COLUMN = 1
FINAL_STATUS_COLUMN = 11
STATUS_COLUMN = 17


class AIAnalyzer:
    """AI-powered analysis engine for validation and recommendations"""
    
    def __init__(self):
        self.analysis_results = []
    
    def analyze_evaluation_data(self, eval_data):
        """
        Analyze evaluation data for quality and completeness
        
        Args:
            eval_data: Dictionary of evaluation results by OpCode
            
        Returns:
            dict: Analysis results with recommendations
        """
        analysis = {
            'total_evaluations': len(eval_data),
            'status_distribution': defaultdict(int),
            'recommendations': [],
            'warnings': [],
            'quality_score': 100
        }
        
        # Count status distribution
        for opcode, results in eval_data.items():
            for result in results:
                status = result.get('status', 'N/A')
                analysis['status_distribution'][status] += 1
        
        # Calculate quality metrics
        red_count = analysis['status_distribution'].get('RED', 0)
        total_count = sum(analysis['status_distribution'].values())
        
        if total_count > 0:
            failure_rate = (red_count / total_count) * 100
            
            # Generate recommendations based on failure rate
            if failure_rate > 50:
                analysis['recommendations'].append(
                    f"⚠️ High failure rate detected ({failure_rate:.1f}%). "
                    "Consider reviewing test procedures or requirements."
                )
                analysis['quality_score'] -= 30
            elif failure_rate > 30:
                analysis['recommendations'].append(
                    f"⚡ Moderate failure rate ({failure_rate:.1f}%). "
                    "Some operations may need attention."
                )
                analysis['quality_score'] -= 15
            else:
                analysis['recommendations'].append(
                    f"✓ Good test performance ({100-failure_rate:.1f}% pass rate)."
                )
        
        # Check for missing evaluations
        if analysis['total_evaluations'] < 20:
            analysis['warnings'].append(
                "📊 Limited evaluation data. Consider testing more operations."
            )
            analysis['quality_score'] -= 10
        
        return analysis
    
    def validate_heatmap_structure(self, heatmap_sheet):
        """
        Validate the HeatMap sheet structure
        
        Args:
            heatmap_sheet: The HeatMap worksheet
            
        Returns:
            dict: Validation results
        """
        validation = {
            'valid': True,
            'errors': [],
            'warnings': []
        }
        
        # Check if required columns exist
        try:
            header_row = list(heatmap_sheet.iter_rows(min_row=2, max_row=2, values_only=True))[0]
            if len(header_row) < 18:
                validation['errors'].append(
                    "HeatMap sheet structure invalid: Missing required columns"
                )
                validation['valid'] = False
        except Exception as e:
            validation['errors'].append(f"Error reading HeatMap structure: {str(e)}")
            validation['valid'] = False
        
        return validation
    
    def generate_insights(self, update_results):
        """
        Generate intelligent insights from update results
        
        Args:
            update_results: Results from the update operation
            
        Returns:
            list: List of insight strings
        """
        insights = []
        
        sub_ops = update_results.get('sub_operations_updated', 0)
        parent_ops = update_results.get('parent_operations_updated', 0)
        
        if sub_ops > 0 and parent_ops > 0:
            insights.append(
                f"✓ Successfully updated {sub_ops} sub-operations and "
                f"{parent_ops} parent operations"
            )
        
        if parent_ops > 0:
            insights.append(
                "💡 Parent operation statuses were automatically calculated "
                "from sub-operation results"
            )
        
        # Add recommendations for next steps
        insights.append(
            "📋 Next steps: Review the updated HeatMap sheet and verify all statuses"
        )
        
        return insights


class HeatMapUpdaterGUI:
    """Main GUI application for HeatMap updates"""
    
    def __init__(self, root):
        self.root = root
        self.root.title("AVLDrive HeatMap Updater - AI-Powered")
        self.root.geometry("900x700")
        
        # Initialize AI analyzer
        self.ai_analyzer = AIAnalyzer()
        
        # Variables
        self.file_path = tk.StringVar()
        self.output_path = tk.StringVar()
        
        # Setup UI
        self.setup_ui()
        
    def setup_ui(self):
        """Setup the user interface"""
        
        # Header
        header_frame = ttk.Frame(self.root, padding="10")
        header_frame.pack(fill=tk.X)
        
        title_label = ttk.Label(
            header_frame, 
            text="AVLDrive HeatMap Updater with AI",
            font=("Arial", 16, "bold")
        )
        title_label.pack()
        
        subtitle_label = ttk.Label(
            header_frame,
            text="Intelligent automation for HeatMap status updates",
            font=("Arial", 10, "italic")
        )
        subtitle_label.pack()
        
        # File selection frame
        file_frame = ttk.LabelFrame(self.root, text="File Selection", padding="10")
        file_frame.pack(fill=tk.X, padx=10, pady=5)
        
        # Input file
        ttk.Label(file_frame, text="Excel File:").grid(row=0, column=0, sticky=tk.W, pady=5)
        ttk.Entry(file_frame, textvariable=self.file_path, width=60).grid(row=0, column=1, padx=5)
        ttk.Button(file_frame, text="Browse", command=self.browse_file).grid(row=0, column=2)
        
        # Output file
        ttk.Label(file_frame, text="Output File:").grid(row=1, column=0, sticky=tk.W, pady=5)
        ttk.Entry(file_frame, textvariable=self.output_path, width=60).grid(row=1, column=1, padx=5)
        ttk.Button(file_frame, text="Browse", command=self.browse_output).grid(row=1, column=2)
        
        # Action buttons frame
        button_frame = ttk.Frame(self.root, padding="10")
        button_frame.pack(fill=tk.X)
        
        self.analyze_btn = ttk.Button(
            button_frame, 
            text="🔍 Analyze File", 
            command=self.analyze_file,
            width=20
        )
        self.analyze_btn.pack(side=tk.LEFT, padx=5)
        
        self.update_btn = ttk.Button(
            button_frame,
            text="▶ Update HeatMap",
            command=self.update_heatmap,
            width=20,
            state=tk.DISABLED
        )
        self.update_btn.pack(side=tk.LEFT, padx=5)
        
        self.export_btn = ttk.Button(
            button_frame,
            text="📊 Export Report",
            command=self.export_report,
            width=20,
            state=tk.DISABLED
        )
        self.export_btn.pack(side=tk.LEFT, padx=5)
        
        # Progress frame
        progress_frame = ttk.LabelFrame(self.root, text="Progress", padding="10")
        progress_frame.pack(fill=tk.X, padx=10, pady=5)
        
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(
            progress_frame,
            variable=self.progress_var,
            maximum=100
        )
        self.progress_bar.pack(fill=tk.X, pady=5)
        
        self.status_label = ttk.Label(progress_frame, text="Ready", foreground="green")
        self.status_label.pack()
        
        # Results frame with tabs
        results_frame = ttk.LabelFrame(self.root, text="Results & AI Analysis", padding="10")
        results_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        # Create notebook for tabs
        self.notebook = ttk.Notebook(results_frame)
        self.notebook.pack(fill=tk.BOTH, expand=True)
        
        # Log tab
        log_frame = ttk.Frame(self.notebook)
        self.notebook.add(log_frame, text="📝 Log")
        
        self.log_text = scrolledtext.ScrolledText(log_frame, height=15, width=80)
        self.log_text.pack(fill=tk.BOTH, expand=True)
        
        # AI Insights tab
        insights_frame = ttk.Frame(self.notebook)
        self.notebook.add(insights_frame, text="🤖 AI Insights")
        
        self.insights_text = scrolledtext.ScrolledText(insights_frame, height=15, width=80)
        self.insights_text.pack(fill=tk.BOTH, expand=True)
        
        # Statistics tab
        stats_frame = ttk.Frame(self.notebook)
        self.notebook.add(stats_frame, text="📊 Statistics")
        
        self.stats_text = scrolledtext.ScrolledText(stats_frame, height=15, width=80)
        self.stats_text.pack(fill=tk.BOTH, expand=True)
        
        # Initial log message
        self.log_message("Application started. Select an Excel file to begin.", "INFO")
        
    def browse_file(self):
        """Browse for input Excel file"""
        filename = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=[
                ("Excel Macro Files", "*.xlsm"),
                ("Excel Files", "*.xlsx"),
                ("All Files", "*.*")
            ]
        )
        if filename:
            self.file_path.set(filename)
            # Auto-set output path
            base, ext = os.path.splitext(filename)
            self.output_path.set(f"{base}_updated{ext}")
            self.log_message(f"Selected file: {filename}", "INFO")
    
    def browse_output(self):
        """Browse for output file location"""
        filename = filedialog.asksaveasfilename(
            title="Save Output File As",
            filetypes=[
                ("Excel Macro Files", "*.xlsm"),
                ("Excel Files", "*.xlsx"),
                ("All Files", "*.*")
            ],
            defaultextension=".xlsm"
        )
        if filename:
            self.output_path.set(filename)
            self.log_message(f"Output will be saved to: {filename}", "INFO")
    
    def log_message(self, message, level="INFO"):
        """Add message to log with timestamp"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        formatted_message = f"[{timestamp}] [{level}] {message}\n"
        
        self.log_text.insert(tk.END, formatted_message)
        self.log_text.see(tk.END)
        
        # Color code by level
        if level == "ERROR":
            self.log_text.tag_add("error", "end-2l", "end-1l")
            self.log_text.tag_config("error", foreground="red")
        elif level == "SUCCESS":
            self.log_text.tag_add("success", "end-2l", "end-1l")
            self.log_text.tag_config("success", foreground="green")
        elif level == "WARNING":
            self.log_text.tag_add("warning", "end-2l", "end-1l")
            self.log_text.tag_config("warning", foreground="orange")
    
    def update_status(self, message, color="black"):
        """Update status label"""
        self.status_label.config(text=message, foreground=color)
        self.root.update_idletasks()
    
    def analyze_file(self):
        """Analyze the selected Excel file using AI"""
        if not self.file_path.get():
            messagebox.showerror("Error", "Please select an Excel file first")
            return
        
        # Run analysis in background thread
        thread = threading.Thread(target=self._analyze_file_thread)
        thread.daemon = True
        thread.start()
    
    def _analyze_file_thread(self):
        """Background thread for file analysis"""
        try:
            self.update_status("Analyzing file...", "blue")
            self.progress_var.set(10)
            self.log_message("Starting AI analysis...", "INFO")
            
            # Load workbook
            wb = load_workbook(self.file_path.get(), keep_vba=True, data_only=False)
            self.progress_var.set(30)
            
            # Analyze Evaluation Results
            self.log_message("Reading Evaluation Results sheet...", "INFO")
            eval_sheet = wb["Evaluation Results"]
            
            eval_data = defaultdict(list)
            for row in eval_sheet.iter_rows(min_row=2, max_row=eval_sheet.max_row, values_only=True):
                opcode = self._normalize_opcode(row[OPCODE_COLUMN])
                final_status = row[FINAL_STATUS_COLUMN]
                
                if opcode and final_status and final_status not in ["N/A", None]:
                    eval_data[opcode].append({
                        'operation': row[OPERATION_COLUMN],
                        'status': final_status
                    })
            
            self.progress_var.set(60)
            
            # Run AI analysis
            self.log_message("Running AI analysis on evaluation data...", "INFO")
            analysis = self.ai_analyzer.analyze_evaluation_data(eval_data)
            
            # Validate HeatMap structure
            self.log_message("Validating HeatMap sheet structure...", "INFO")
            heatmap_sheet = wb["HeatMap Sheet"]
            validation = self.ai_analyzer.validate_heatmap_structure(heatmap_sheet)
            
            self.progress_var.set(90)
            
            # Display results
            self._display_analysis_results(analysis, validation)
            
            wb.close()
            
            self.progress_var.set(100)
            self.update_status("Analysis complete", "green")
            self.log_message("AI analysis completed successfully", "SUCCESS")
            
            # Enable update button
            self.update_btn.config(state=tk.NORMAL)
            
        except Exception as e:
            self.log_message(f"Error during analysis: {str(e)}", "ERROR")
            self.update_status("Analysis failed", "red")
            messagebox.showerror("Analysis Error", f"Failed to analyze file:\n{str(e)}")
        finally:
            self.progress_var.set(0)
    
    def _display_analysis_results(self, analysis, validation):
        """Display AI analysis results"""
        # Clear previous insights
        self.insights_text.delete(1.0, tk.END)
        self.stats_text.delete(1.0, tk.END)
        
        # Display insights
        self.insights_text.insert(tk.END, "=== AI ANALYSIS RESULTS ===\n\n")
        self.insights_text.insert(tk.END, f"Quality Score: {analysis['quality_score']}/100\n\n")
        
        if analysis['recommendations']:
            self.insights_text.insert(tk.END, "Recommendations:\n")
            for rec in analysis['recommendations']:
                self.insights_text.insert(tk.END, f"  {rec}\n")
            self.insights_text.insert(tk.END, "\n")
        
        if analysis['warnings']:
            self.insights_text.insert(tk.END, "Warnings:\n")
            for warn in analysis['warnings']:
                self.insights_text.insert(tk.END, f"  {warn}\n")
            self.insights_text.insert(tk.END, "\n")
        
        if validation['errors']:
            self.insights_text.insert(tk.END, "Validation Errors:\n")
            for error in validation['errors']:
                self.insights_text.insert(tk.END, f"  ❌ {error}\n")
            self.insights_text.insert(tk.END, "\n")
        
        # Display statistics
        self.stats_text.insert(tk.END, "=== STATISTICS ===\n\n")
        self.stats_text.insert(tk.END, f"Total Evaluations: {analysis['total_evaluations']}\n\n")
        self.stats_text.insert(tk.END, "Status Distribution:\n")
        
        for status, count in analysis['status_distribution'].items():
            percentage = (count / sum(analysis['status_distribution'].values())) * 100
            self.stats_text.insert(tk.END, f"  {status}: {count} ({percentage:.1f}%)\n")
    
    def update_heatmap(self):
        """Update HeatMap with evaluation results"""
        if not self.file_path.get():
            messagebox.showerror("Error", "Please select an Excel file first")
            return
        
        if not self.output_path.get():
            messagebox.showerror("Error", "Please specify output file location")
            return
        
        # Confirm action
        if not messagebox.askyesno(
            "Confirm Update",
            "This will update the HeatMap sheet with evaluation results. Continue?"
        ):
            return
        
        # Run update in background thread
        thread = threading.Thread(target=self._update_heatmap_thread)
        thread.daemon = True
        thread.start()
    
    def _update_heatmap_thread(self):
        """Background thread for HeatMap update"""
        try:
            self.update_status("Updating HeatMap...", "blue")
            self.progress_var.set(5)
            self.log_message("Starting HeatMap update...", "INFO")
            
            # Load workbook
            wb = load_workbook(self.file_path.get(), keep_vba=True)
            self.progress_var.set(10)
            
            # Read Evaluation Results
            self.log_message("Reading evaluation results...", "INFO")
            eval_sheet = wb["Evaluation Results"]
            eval_results = defaultdict(list)
            
            for row in eval_sheet.iter_rows(min_row=2, max_row=eval_sheet.max_row, values_only=True):
                opcode = self._normalize_opcode(row[OPCODE_COLUMN])
                final_status = row[FINAL_STATUS_COLUMN]
                
                if opcode and final_status and final_status not in ["N/A", None]:
                    eval_results[opcode].append({
                        'operation': row[OPERATION_COLUMN],
                        'status': final_status
                    })
            
            self.progress_var.set(30)
            self.log_message(f"Found {len(eval_results)} OpCodes with evaluation data", "INFO")
            
            # Update HeatMap Sheet
            self.log_message("Updating HeatMap Sheet...", "INFO")
            heatmap_sheet = wb["HeatMap Sheet"]
            
            parent_operations = {}
            sub_operations_by_parent = defaultdict(list)
            sub_op_count = 0
            
            # First pass: Update sub-operations
            total_rows = heatmap_sheet.max_row - 3
            for idx, row_idx in enumerate(range(4, heatmap_sheet.max_row + 1)):
                row = list(heatmap_sheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=False))[0]
                opcode = self._normalize_opcode(row[OPCODE_COLUMN].value)
                status_cell = row[STATUS_COLUMN]
                
                if not opcode:
                    continue
                
                # Update progress
                progress = 30 + (idx / total_rows) * 40
                self.progress_var.set(progress)
                
                if self._is_parent_operation(opcode):
                    parent_operations[row_idx] = opcode
                else:
                    if opcode in eval_results:
                        statuses = [er['status'] for er in eval_results[opcode]]
                        status = self._get_worst_status(statuses)
                        
                        if status != "N/A":
                            status_cell.value = "●"
                            color = self._get_color_for_status(status)
                            if color:
                                status_cell.font = Font(color=color, size=14)
                            sub_op_count += 1
                            
                            parent_opcode = self._get_parent_opcode(opcode)
                            if parent_opcode:
                                sub_operations_by_parent[parent_opcode].append(status)
            
            self.progress_var.set(70)
            self.log_message(f"Updated {sub_op_count} sub-operations", "INFO")
            
            # Second pass: Update parent operations
            parent_op_count = 0
            for row_idx, parent_opcode in parent_operations.items():
                row = list(heatmap_sheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=False))[0]
                status_cell = row[STATUS_COLUMN]
                
                sub_statuses = sub_operations_by_parent.get(parent_opcode, [])
                
                if sub_statuses:
                    worst_status = self._get_worst_status(sub_statuses)
                    status_text = self._get_text_for_status(worst_status)
                    
                    if status_text:
                        status_cell.value = status_text
                        color = self._get_color_for_status(worst_status)
                        if color:
                            status_cell.font = Font(color=color, bold=True, size=11)
                        parent_op_count += 1
            
            self.progress_var.set(90)
            self.log_message(f"Updated {parent_op_count} parent operations", "INFO")
            
            # Save workbook
            self.log_message("Saving updated workbook...", "INFO")
            wb.save(self.output_path.get())
            wb.close()
            
            self.progress_var.set(100)
            
            # Generate AI insights on update results
            update_results = {
                'sub_operations_updated': sub_op_count,
                'parent_operations_updated': parent_op_count
            }
            insights = self.ai_analyzer.generate_insights(update_results)
            
            # Display insights
            self.insights_text.insert(tk.END, "\n=== UPDATE INSIGHTS ===\n\n")
            for insight in insights:
                self.insights_text.insert(tk.END, f"{insight}\n")
            
            self.update_status("Update complete", "green")
            self.log_message("HeatMap update completed successfully!", "SUCCESS")
            self.log_message(f"Output saved to: {self.output_path.get()}", "SUCCESS")
            
            # Enable export button
            self.export_btn.config(state=tk.NORMAL)
            
            messagebox.showinfo(
                "Success",
                f"HeatMap updated successfully!\n\n"
                f"Sub-operations updated: {sub_op_count}\n"
                f"Parent operations updated: {parent_op_count}\n\n"
                f"File saved to:\n{self.output_path.get()}"
            )
            
        except Exception as e:
            self.log_message(f"Error during update: {str(e)}", "ERROR")
            self.update_status("Update failed", "red")
            messagebox.showerror("Update Error", f"Failed to update HeatMap:\n{str(e)}")
        finally:
            self.progress_var.set(0)
    
    def export_report(self):
        """Export analysis report to file"""
        try:
            report_path = filedialog.asksaveasfilename(
                title="Export Report",
                filetypes=[("Text Files", "*.txt"), ("JSON Files", "*.json")],
                defaultextension=".txt"
            )
            
            if report_path:
                with open(report_path, 'w') as f:
                    f.write("=== AVLDrive HeatMap Updater - Report ===\n\n")
                    f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
                    f.write("=== Log ===\n")
                    f.write(self.log_text.get(1.0, tk.END))
                    f.write("\n=== AI Insights ===\n")
                    f.write(self.insights_text.get(1.0, tk.END))
                    f.write("\n=== Statistics ===\n")
                    f.write(self.stats_text.get(1.0, tk.END))
                
                self.log_message(f"Report exported to: {report_path}", "SUCCESS")
                messagebox.showinfo("Success", f"Report exported to:\n{report_path}")
        
        except Exception as e:
            self.log_message(f"Error exporting report: {str(e)}", "ERROR")
            messagebox.showerror("Export Error", f"Failed to export report:\n{str(e)}")
    
    # Helper methods
    def _normalize_opcode(self, opcode):
        """Normalize OpCode to string format"""
        try:
            if opcode is None:
                return None
            return str(int(float(opcode)))
        except (ValueError, TypeError):
            return None
    
    def _is_parent_operation(self, opcode):
        """Check if OpCode is a parent operation"""
        opcode_str = str(opcode)
        if not opcode_str.isdigit():
            return False
        trailing_zeros = len(opcode_str) - len(opcode_str.rstrip('0'))
        return trailing_zeros >= 4
    
    def _get_parent_opcode(self, opcode_str):
        """Get parent OpCode for a sub-operation"""
        if not opcode_str.isdigit() or len(opcode_str) < 8:
            return None
        return opcode_str[:4] + "0000"
    
    def _get_color_for_status(self, status):
        """Get color code for status"""
        if status == "RED":
            return RED_COLOR
        elif status == "GREEN":
            return GREEN_COLOR
        elif status == "YELLOW":
            return YELLOW_COLOR
        return None
    
    def _get_text_for_status(self, status):
        """Get text for parent operation status"""
        if status == "RED":
            return "NOK"
        elif status == "GREEN":
            return "OK"
        elif status == "YELLOW":
            return "acceptable"
        return None
    
    def _get_worst_status(self, statuses):
        """Get worst status from list"""
        if "RED" in statuses:
            return "RED"
        elif "YELLOW" in statuses:
            return "YELLOW"
        elif "GREEN" in statuses:
            return "GREEN"
        return "N/A"


def main():
    """Main entry point"""
    root = tk.Tk()
    app = HeatMapUpdaterGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()
